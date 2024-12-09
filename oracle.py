import os  # Módulo para interagir com o sistema operacional, como manipulação de arquivos e diretórios
import re  # Módulo para trabalhar com expressões regulares
import logging  # Módulo para registrar eventos e erros
from datetime import datetime  # Módulo para manipulação de datas e horas
import pandas as pd  # Biblioteca para manipulação e análise de dados
import openpyxl  # Biblioteca para manipulação de arquivos Excel
import oracledb  # Módulo para conectar-se ao banco de dados Oracle
from openpyxl import Workbook  # Classe para criar novos arquivos Excel
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side  # Estilos para formatação de células no Excel
import configparser  # Módulo para ler arquivos de configuração

# Configuração do logging: define onde os logs serão salvos, o nível de severidade e o formato das mensagens
logging.basicConfig(
    filename=(r'\\planilha_log.log'), 
    level=logging.INFO, # Define o nível de log como INFO
    format='%(asctime)s - %(levelname)s - %(message)s', # Define o formato das mensagens de log
    filemode='w'
    )

# Estilo de preenchimento para o cabeçalho da planilha Excel
header_fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')  
# Estilo de borda para as células da planilha Excel
border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))  

def collect_pdfs(directories):
    # Função para coletar arquivos PDF de uma lista de diretórios
    pdf_files = []  # Lista que armazenará os nomes dos arquivos PDF encontrados
    pdf_sources = []  # Lista que armazenará os diretórios de origem dos arquivos PDF encontrados
    
    for directory in directories:  # Itera sobre cada diretório fornecido
        # Verifica se o diretório existe
        if not os.path.isdir(directory):
            logging.error(f"Diretório não encontrado: {directory}")  # Registra um erro se o diretório não existe
            continue  # Passa para o próximo diretório
        
        files_in_directory = os.listdir(directory)  # Lista todos os arquivos no diretório
        
        # Verifica se o diretório está vazio
        if not files_in_directory:
            logging.warning(f"O diretório {directory} está vazio.")  # Registra um aviso se o diretório estiver vazio
            continue  # Passa para o próximo diretório

        for file in files_in_directory:  # Itera sobre cada arquivo encontrado no diretório
            # Verifica se o arquivo termina com a extensão PDF
            if file.endswith('.pdf'):
                cleaned_uc = re.sub(r'\D', '', file)  # Remove todos os caracteres não numéricos do nome do arquivo
                pdf_files.append(cleaned_uc)  # Adiciona o nome do arquivo (apenas números) à lista de arquivos PDF
                pdf_sources.append(directory)  # Adiciona o diretório à lista de fontes de arquivos PDF
                logging.info(f"Arquivo PDF encontrado: {file} no diretório {directory}")  # Registra a informação do arquivo encontrado
            else:
                logging.warning(f"Arquivo ignorado (não é PDF): {file}")  # Registra um aviso se o arquivo não for um PDF

    return pdf_files, pdf_sources  # Retorna as listas de arquivos PDF e suas fontes

def get_current_date(date_format="%d.%m.%Y"):
    # Função para obter a data atual formatada
    return datetime.now().strftime(date_format)  # Retorna a data atual no formato especificado

def connect_to_database(config, retries=3):
    # Função para conectar ao banco de dados Oracle
    dsn = oracledb.makedsn(config['address'], config['port'], service_name=config['service_name'])  # Cria o DSN para a conexão com o banco de dados
    for attempt in range(retries):  # Tenta conectar ao banco de dados até o número de tentativas especificado
        try:
            return oracledb.connect(user=config['user'], password=config['password'], dsn=dsn)  # Tenta estabelecer a conexão com as credenciais fornecidas
        except oracledb.DatabaseError as e:  # Captura erros de banco de dados
            error_code = e.args[0].code  # Obtém o código do erro
            logging.error(f"Erro ao conectar ao banco de dados: {e}. Tentativa {attempt + 1} de {retries}.")  # Registra um erro de conexão, incluindo a tentativa atual
            if attempt == retries - 1:  # Se for a última tentativa
                if error_code in [1010, 28000]:  # Verifica se o erro é de usuário ou senha inválidos
                    logging.warning("Usuário ou senha inválidos. O acesso ao SQL será bloqueado após 3 tentativas.")  # Registra aviso
                    print("Usuário ou senha inválidos. O acesso ao SQL será bloqueado após 3 tentativas.")  # Informa ao usuário
                raise  # Levanta a exceção para ser tratada na chamada da função

def execute_query(cursor, ucs, pdf_files):
    # Função para executar uma consulta SQL no banco de dados
    sql = f"""
        SELECT DISTINCT
            a.cod_un_cons_uee,  
            a.cod_loc_uee  
        FROM 
            rededes.cad_uc_ee a  
        WHERE
            a.cod_un_cons_uee IN {ucs}  
    """
    try:
        cursor.execute(sql)  # Executa a consulta SQL
        results = cursor.fetchall()  # Obtém todos os resultados da consulta
        logging.info(f"Resultados da consulta SQL: {results}")  # Registra os resultados obtidos

        found_ucs = {str(row[0]) for row in results}  # Cria um conjunto com os códigos encontrados na consulta

        # Verifica se todos os arquivos PDF estão presentes nos resultados
        for pdf in pdf_files:
            if pdf not in found_ucs:  # Se um PDF não foi encontrado nos resultados
                logging.warning(f"UC não encontrada na consulta: {pdf}")  # Registra um aviso

        return results if isinstance(results, list) else []  # Retorna os resultados se forem uma lista, caso contrário retorna uma lista vazia
    except oracledb.DatabaseError as e:  # Captura erros de execução da consulta
        logging.error(f"Erro ao executar a consulta SQL: {e}")  # Registra o erro
        return []  # Retorna uma lista vazia em caso de erro

def load_localities(file_path):
    # Função para carregar localidades de um arquivo de texto
    localities = {}  # Dicionário para armazenar localidades
    try:
        with open(file_path, 'r') as file:  # Abre o arquivo de localidades para leitura
            for line in file:  # Itera sobre cada linha do arquivo
                cod_loc, locality = line.strip().split(',')  # Divide a linha em código e nome da localidade
                localities[cod_loc] = locality  # Adiciona ao dicionário
    except Exception as e:  # Captura qualquer erro durante a leitura do arquivo
        logging.error(f"Erro ao carregar localidades do arquivo: {e}")  # Registra o erro
    return localities  # Retorna o dicionário de localidades

def format_dataframe(results, current_date, localities, pdf_files, pdf_sources):
    # Função para formatar os resultados da consulta em um DataFrame do pandas
    results_df = pd.DataFrame(results, columns=['cod_un_cons_uee', 'cod_loc_uee'])  # Cria um DataFrame com os resultados
    results_df['Data'] = current_date  # Adiciona a data atual ao DataFrame
    results_df['cod_loc_uee'] = results_df['cod_loc_uee'].astype(str).str.lstrip('0')  # Remove zeros à esquerda dos códigos de localidade

    # Mapeia os códigos de localidade para seus nomes correspondentes
    results_df['Localidade'] = results_df['cod_loc_uee'].map(localities).fillna("Localidade não encontrada")  # Preenche com aviso se não encontrado

    # Mapeia os códigos de unidades consumidoras para suas fontes de PDF
    results_df['PDF'] = results_df['cod_un_cons_uee'].astype(str).map(dict(zip(pdf_files, pdf_sources)))  # Cria coluna de PDF

    # Verifica se a fonte do PDF corresponde ao diretório ENERGEC
    energenc_directory = (r'S:\SRC\01_Gestao_da_Receita\01-Recuperacao_Energia\03-Usuarios\ALEX GUIDONI\--- ANÁLISE TOIS FTP\===ENERGEC')
    results_df['Localidade'] = results_df.apply(
        lambda row: f"ENERGEC - {row['Localidade']}" if row['PDF'] == energenc_directory else row['Localidade'],  # Adiciona prefixo 'ENERGEC' se a fonte do PDF for do diretório ENERGEC
        axis=1  # Aplica a função ao longo das linhas
    )

    # Renomeia as colunas do DataFrame para nomes mais amigáveis
    results_df.rename(columns={
        'cod_un_cons_uee': 'UC',  # Renomeia 'cod_un_cons_uee' para 'UC'
        'Localidade': 'UTD',  # Renomeia 'Localidade' para 'UTD'
        'Data': 'DATA'  # Renomeia 'Data' para 'DATA'
    }, inplace=True)

    # Reorganiza as colunas do DataFrame na ordem desejada
    results_df = results_df[['UC', 'UTD', 'DATA']]
    # Formata a coluna 'DATA' para o formato 'dd/mm/yyyy'
    results_df['DATA'] = pd.to_datetime(results_df['DATA'], format='%d.%m.%Y').dt.strftime('%d/%m/%Y')

    return results_df  # Retorna o DataFrame formatado

def save_to_excel(results_df, current_date):
    # Função para salvar o DataFrame em um arquivo Excel
    if results_df.empty:  # Verifica se o DataFrame está vazio
        logging.warning("O DataFrame gerado está vazio. Nenhum dado a ser salvo.")  # Registra um aviso
        return  # Sai da função se não houver dados

    # Define o diretório onde o arquivo Excel será salvo
    save_directory = (r'//destiny//directory')
    # Cria o nome do arquivo Excel com a data atual
    spreadsheet_name = os.path.join(save_directory, f"Processos retirados do FTP em {current_date}.xlsx")
    
    # Salva o DataFrame em um arquivo Excel
    results_df.to_excel(spreadsheet_name, index=False, sheet_name='Processos_Retirados')  # Salva sem índice
    logging.info(f"Planilha salva com sucesso: {spreadsheet_name}")  # Registra a informação de que a planilha foi salva

    wb = openpyxl.load_workbook(spreadsheet_name)  # Carrega o arquivo Excel salvo
    ws = wb.active  # Seleciona a planilha ativa
    
    # Ajusta as larguras das colunas para melhor visualização
    ws.column_dimensions['A'].width = 11  # Largura da coluna UC
    ws.column_dimensions['B'].width = 31  # Largura da coluna UTD
    ws.column_dimensions['C'].width = 11  # Largura da coluna DATA

    # Formata as células da planilha
    for row in range(1, ws.max_row + 1):  # Itera sobre todas as linhas
        for column in range(1, ws.max_column + 1):  # Itera sobre todas as colunas
            cell = ws.cell(row=row, column=column)  # Seleciona a célula atual
            cell.alignment = Alignment(horizontal='left', vertical='center')  # Alinha o texto à esquerda e centraliza verticalmente
            cell.border = border_style  # Aplica o estilo de borda à célula
            if row == 1:  # Se for a primeira linha (cabeçalho)
                cell.font = Font(bold=True, color='000000')  # Aplica negrito e cor ao texto do cabeçalho
                cell.fill = header_fill  # Aplica o preenchimento do cabeçalho

    wb.save(spreadsheet_name)  # Salva as alterações feitas na planilha
    wb.close()  # Fecha o arquivo Excel
    print(f"Planilha criada: {spreadsheet_name}")  # Informa ao usuário que a planilha foi criada

def main():
    # Função principal que controla o fluxo do programa
    directories = [  # Lista de diretórios a serem verificados em busca de arquivos PDF
        (r'//source//directory//1'),
        (r'//source//directory//2'),
        (r'//source//directory//3'),
        (r'//source//directory//4'),
        (r'//source//directory//5')
    ]

    # Caminho do arquivo que contém as localidades
    localities_file_path = (r'\\Localidades.txt')
    localities = load_localities(localities_file_path)  # Carrega as localidades do arquivo

    pdf_files, pdf_sources = collect_pdfs(directories)  # Coleta os arquivos PDF e suas fontes

    if not pdf_files:  # Verifica se nenhum arquivo PDF foi encontrado
        logging.warning("Nenhum arquivo PDF encontrado.")  # Registra um aviso
        print("Nenhum arquivo PDF encontrado.")  # Informa ao usuário
        return  # Sai da função principal

    # Formata os códigos das unidades consumidoras para a consulta
    ucs = "','".join(pdf_files)  # Junta os códigos em uma string formatada
    ucs = f"('{ucs}')"  # Adiciona parênteses para a consulta SQL

    current_date = get_current_date()  # Obtém a data atual

    config = configparser.ConfigParser()  # Cria um objeto de configuração

    try:
        config.read(r'\\login_sql.ini')  # Lê o arquivo de configuração
        if not config.has_section('database'):  # Verifica se a seção 'database' existe
            raise ValueError("A seção 'database' não foi encontrada no arquivo de configuração.")  # Levanta um erro se não existir

        # Verifica se todas as chaves necessárias estão presentes na seção 'database'
        required_keys = ['user', 'password', 'address', 'port', 'service_name']
        for key in required_keys:
            if key not in config['database']:
                raise ValueError(f"A chave '{key}' não foi encontrada na seção 'database'.")  # Levanta um erro se faltar alguma chave

        # Cria um dicionário com as configurações do banco de dados
        db_config = {
            'user': config['database']['user'],
            'password': config['database']['password'],
            'address': config['database']['address'],
            'port': config['database']['port'],
            'service_name': config['database']['service_name']
        }

    except Exception as e:  # Captura erros ao ler o arquivo de configuração
        logging.error(f"Erro ao ler o arquivo de configuração: {e}")  # Registra o erro
        print("Erro ao ler o arquivo de configuração. Verifique os logs para mais detalhes.")  # Informa ao usuário
        return  # Sai da função principal
    
    try:
        # Tenta conectar ao banco de dados e executar a consulta
        with connect_to_database(db_config) as conn:  # Conecta ao banco de dados
            with conn.cursor() as cursor:  # Cria um cursor para executar consultas
                results = execute_query(cursor, ucs, pdf_files)  # Executa a consulta e obtém os resultados
                logging.info(f"Número de resultados retornados: {len(results)}")  # Registra o número de resultados
                print(f"Número de resultados retornados: {len(results)}")  # Informa ao usuário

        # Formata os resultados em um DataFrame
        results_df = format_dataframe(results, current_date, localities, pdf_files, pdf_sources)
        save_to_excel(results_df, current_date)  # Salva os resultados em um arquivo Excel

    except oracledb.DatabaseError as e:  # Captura erros de banco de dados
        logging.error(f"Erro ao conectar ou executar a consulta no banco de dados: {e}")  # Registra o erro
        print("Ocorreu um erro ao tentar acessar o banco de dados. Verifique os logs para mais detalhes.")  # Informa ao usuário
    except Exception as e:  # Captura erros inesperados
        logging.error(f"Erro inesperado: {e}")  # Registra o erro
        print("Ocorreu um erro inesperado. Verifique os logs para mais detalhes.")  # Informa ao usuário

# Verifica se o script está sendo executado diretamente
if __name__ == "__main__":
    main()  # Chama a função principal
