import openpyxl  # Importa a biblioteca openpyxl para manipulação de arquivos Excel
import os  # Importa a biblioteca os para interagir com o sistema de arquivos
import pandas as pd  # Importa a biblioteca pandas para manipulação de dados
from datetime import datetime  # Importa a classe datetime para trabalhar com datas
import re  # Importa a biblioteca re para expressões regulares
import cx_Oracle  # Importa a biblioteca cx_Oracle para conexão com bancos de dados Oracle
import logging  # Importa a biblioteca logging para registrar logs de eventos
from openpyxl import Workbook  # Importa a classe Workbook da biblioteca openpyxl
from openpyxl.styles import Font, Alignment  # Importa classes para estilização de células
from openpyxl.utils import get_column_letter  # Importa função para obter letras de colunas

# Inicializa o cliente Oracle com o diretório do Instant Client
cx_Oracle.init_oracle_client(lib_dir=r"//sql//instantclient")

# Configura o logging para registrar informações em um arquivo
logging.basicConfig(filename='pdf_collection.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def collect_pdfs(directories):
    pdf_files = []  # Lista para armazenar os arquivos PDF encontrados
    for directory in directories:
        try:
            logging.info(f"Processando diretório: {directory}")  # Registra o diretório em processamento
            if not os.path.isdir(directory):  # Verifica se o diretório existe
                logging.error(f"Diretório não encontrado: {directory}")  # Registra erro se não encontrado
                continue

            files_in_directory = os.listdir(directory)  # Lista os arquivos no diretório
            if not files_in_directory:  # Verifica se o diretório está vazio
                logging.warning(f"O diretório {directory} está vazio.")  # Registra aviso se estiver vazio
                continue 

            for file in files_in_directory:  # Itera sobre os arquivos encontrados
                if file.endswith('.pdf'):  # Verifica se o arquivo é um PDF
                    cleaned_uc = re.sub(r'\D', '', file)  # Limpa o nome do arquivo para extrair apenas números
                    pdf_files.append(cleaned_uc)  # Adiciona o arquivo à lista
                    logging.info(f"Arquivo PDF encontrado: {file}")  # Registra o arquivo encontrado
                else:
                    logging.warning(f"Arquivo ignorado (não é PDF): {file}")  # Registra aviso para arquivos não PDF

            logging.info(f"Processamento do diretório {directory} concluído com sucesso.")  # Registra conclusão do processamento
        except FileNotFoundError:
            logging.error(f"Diretório não encontrado: {directory}")  # Registra erro se o diretório não for encontrado
        except Exception as e:
            logging.error(f"Erro ao processar o diretório {directory}: {e}")  # Registra erro genérico
    return pdf_files  # Retorna a lista de arquivos PDF encontrados

def get_current_date(date_format="%d.%m.%Y"):
    return datetime.now().strftime(date_format)  # Retorna a data atual formatada

def connect_to_database(config, retries=3):
    dsn = cx_Oracle.makedsn(config['address'], config['port'], service_name=config['service_name'])  # Cria o DSN para conexão
    for attempt in range(retries):  # Tenta conectar ao banco de dados com retries
        try:
            return cx_Oracle.connect(config['user'], config['password'], dsn)  # Conecta ao banco de dados
        except cx_Oracle.DatabaseError as e:
            logging.error(f"Erro ao conectar ao banco de dados: {e}. Tentativa {attempt + 1} de {retries}.")  # Registra erro de conexão
            if attempt < retries - 1:
                continue  # Tenta novamente se ainda houver tentativas
            else:
                raise  # Lança exceção se todas as tentativas falharem

def execute_query(cursor, ucs):
    sql = f"""
        SELECT DISTINCT
        a.cod_un_cons_uee,
        a.cod_loc_uee
        FROM 
            rededes.cad_uc_ee a
        WHERE
             a.cod_un_cons_uee IN {ucs}
    """  # SQL para buscar dados no banco de dados
    try:
        cursor.execute(sql)  # Executa a consulta SQL
        return cursor.fetchall()  # Retorna todos os resultados da consulta
    except cx_Oracle.DatabaseError as e:
        logging.error(f"Erro ao executar a consulta SQL: {e}")  # Registra erro na execução da consulta SQL
    return []  # Retorna uma lista vazia em caso de erro

def load_localities(file_path):
    localities = {}  # Dicionário para armazenar localidades
    try:
        with open(file_path, 'r') as file:  # Abre o arquivo de localidades
            for line in file:  # Itera sobre cada linha do arquivo
                cod_loc, locality = line.strip().split(',')  # Divide a linha em código e nome da localidade
                localities[cod_loc] = locality  # Adiciona ao dicionário
    except Exception as e:
        logging.error(f"Erro ao carregar localidades do arquivo: {e}")  # Registra erro se houver problema ao carregar o arquivo
    return localities  # Retorna o dicionário de localidades

def save_to_excel(results, current_date, localities):
    if results:  # Verifica se há resultados para salvar
        results_df = pd.DataFrame(results, columns=['cod_un_cons_uee', 'cod_loc_uee'])  # Cria DataFrame a partir dos resultados
        results_df['Data'] = current_date  # Adiciona a data atual ao DataFrame

        results_df['Localidade'] = results_df['cod_loc_uee'].map(localities)  # Mapeia códigos de localidade para nomes

        results_df = results_df[['cod_un_cons_uee', 'cod_loc_uee', 'Localidade', 'Data']]  # Reorganiza as colunas

        spreadsheet_name = f"Processos recuperados do FTP em {current_date}.xlsx"  # Define o nome do arquivo Excel
        results_df.to_excel(spreadsheet_name, index=False, sheet_name='SQL_Results')  # Salva o DataFrame em um arquivo Excel

        wb = Workbook(spreadsheet_name)  # Cria uma nova planilha
        ws = wb.active  # Obtém a planilha ativa

        for column_cells in ws.columns:  # Ajusta a largura das colunas
            length = max(len(str(cell.value)) for cell in column_cells)  # Calcula o comprimento máximo da coluna
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2  # Define a largura da coluna

        for row in range(1, ws.max_row + 1):  # Formata as células da planilha
            for column in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=column)  # Obtém a célula
                cell.alignment = Alignment(horizontal='center', vertical='center')  # Centraliza o texto
                if row == 1:  # Aplica formatação especial para o cabeçalho
                    cell.font = Font(bold=True, color='FFFFFF')  # Define fonte em negrito e branca
                    cell.fill = openpyxl.styles.PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')  # Define cor de fundo

        wb.save(spreadsheet_name)  # Salva a planilha formatada
        print(f"Planilha criada com formatação: {spreadsheet_name}")  # Imprime mensagem de sucesso
    else:
        print("Nenhum resultado encontrado para a consulta SQL.")  # Mensagem se não houver resultados

def main():
    directories = [  # Lista de diretórios a serem processados
        '//source//directory_1',
        '//source//directory_2',
        '//source//directory_3',
        '//source//directory_4'
    ]

    localities_file_path = '//path//to//cod_loc'  # Caminho do arquivo que contém localidades
    localities = load_localities(localities_file_path)  # Carrega localidades do arquivo

    pdf_files = collect_pdfs(directories)  # Coleta arquivos PDF dos diretórios

    ucs = "','".join(pdf_files)  # Cria uma string com os códigos dos arquivos PDF
    ucs = f"('{ucs}')"  # Formata a string para uso na consulta SQL

    date_format = "%d.%m.%Y"  # Define o formato da data
    current_date = get_current_date(date_format)  # Obtém a data atual formatada

    db_config = {  # Configurações para conexão com o banco de dados
        'user': '---',  # Nome de usuário do banco de dados
        'password': '---',  # Senha do banco de dados
        'address': '---',  # Endereço do banco de dados
        'port': 1234,  # Porta do banco de dados
        'service_name': '--' # Nome do serviço do banco de dados
    }
    try:
        with connect_to_database(db_config) as conn:  # Conecta ao banco de dados
            with conn.cursor() as cursor:  # Cria um cursor para executar consultas
                results = execute_query(cursor, ucs)  # Executa a consulta SQL e obtém os resultados

        save_to_excel(results, current_date, localities)  # Salva os resultados em um arquivo Excel

    except cx_Oracle.DatabaseError as e:
        logging.error(f"Erro ao conectar ou executar a consulta no banco de dados: {e}")  # Registra erro de conexão ou consulta
        print("Ocorreu um erro ao tentar acessar o banco de dados. Verifique os logs para mais detalhes.")  # Mensagem de erro para o usuário
    except Exception as e:
        logging.error(f"Erro inesperado: {e}")  # Registra erro inesperado
        print("Ocorreu um erro inesperado. Verifique os logs para mais detalhes.")  # Mensagem de erro para o usuário

if __name__ == "__main__":  # Verifica se o script está sendo executado diretamente
    main()  # Chama a função principal
